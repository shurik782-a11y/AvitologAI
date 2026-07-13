"""Import competitor ads table (CSV/XLSX) → compressed competitor_insights."""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from sqlalchemy.orm import Session

from app.config import settings
from app.db import AppSettings, MetricEvent, Project
from app.schemas import CompetitorsImportResult
from app.services.openrouter import OpenRouterError, chat_completions
from app.services.prompts import COMPETITOR_COMPRESS_SYSTEM

_MAX_ROWS_READ = 2000
_MAX_ROWS_LLM = 40
_DESC_CAP = 280


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _pick(row: dict[str, str], *keys: str) -> str:
    for k in keys:
        for rk, rv in row.items():
            if _norm_header(rk) == k or k in _norm_header(rk):
                return (rv or "").strip()
    return ""


def _parse_csv(raw: bytes) -> list[dict[str, str]]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(r) for r in reader if r]


def _parse_xlsx(raw: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Для XLSX установите openpyxl или сохраните таблицу как CSV") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(c or "").strip() for c in next(rows_iter)]
    except StopIteration:
        return []
    out: list[dict[str, str]] = []
    for row in rows_iter:
        item = {headers[i]: "" if v is None else str(v) for i, v in enumerate(row) if i < len(headers)}
        if any(str(v).strip() for v in item.values()):
            out.append(item)
    return out


def parse_competitors_file(raw: bytes, filename: str) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _parse_xlsx(raw)
    return _parse_csv(raw)


def select_top_rows(rows: list[dict[str, str]], limit: int = _MAX_ROWS_LLM) -> list[dict[str, str]]:
    scored: list[tuple[float, dict[str, str]]] = []
    for row in rows[:_MAX_ROWS_READ]:
        title = _pick(row, "title", "заголовок", "name", "название")
        desc = _pick(row, "description", "описание", "text", "текст")
        views_raw = _pick(row, "views", "просмотры", "view", "hits")
        try:
            views = float(re.sub(r"[^\d.]", "", views_raw) or "0")
        except ValueError:
            views = 0.0
        if not title and not desc:
            continue
        scored.append(
            (
                views,
                {
                    "title": title[:120],
                    "description": desc[:_DESC_CAP],
                    "views": str(int(views)) if views else "",
                },
            )
        )
    scored.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in scored[:limit]]


async def import_competitors_table(
    db: Session,
    project: Project,
    raw: bytes,
    *,
    filename: str,
) -> CompetitorsImportResult:
    rows = parse_competitors_file(raw, filename)
    if not rows:
        raise ValueError("В файле нет строк")
    top = select_top_rows(rows)
    cfg = db.get(AppSettings, 1)
    api_key = (cfg.openrouter_api_key if cfg else "") or settings.openrouter_api_key
    model = (settings.onboarding_model or "openrouter/free").strip()

    compact = "\n".join(
        f"- {r['title']} | views={r['views'] or '—'} | {r['description']}" for r in top
    )
    insights = ""
    if api_key:
        try:
            raw_llm = await chat_completions(
                api_key,
                model=model,
                messages=[
                    {"role": "system", "content": COMPETITOR_COMPRESS_SYSTEM},
                    {
                        "role": "user",
                        "content": f"Ниша/проект: {project.name}\nТема: {project.theme or '—'}\n\nОбъявления:\n{compact}",
                    },
                ],
                temperature=0.2,
                max_tokens=900,
            )
            insights = (
                (raw_llm.get("choices") or [{}])[0].get("message", {}).get("content") or ""
            ).strip()
        except OpenRouterError:
            insights = ""
    if not insights:
        # Offline fallback: titles only
        titles = [r["title"] for r in top if r["title"]][:15]
        insights = "Топ заголовков конкурентов:\n" + "\n".join(f"• {t}" for t in titles)

    project.competitor_insights = insights[:4000]
    db.add(project)
    db.add(
        MetricEvent(
            project_id=project.id,
            name="competitors.import",
            value=float(len(top)),
            payload={"rows_read": len(rows), "rows_used": len(top)},
        )
    )
    db.commit()
    db.refresh(project)
    return CompetitorsImportResult(
        rows_read=len(rows),
        rows_used=len(top),
        competitor_insights=project.competitor_insights,
        message=f"Импортировано: прочитано {len(rows)}, в анализ {len(top)}. Insights сохранены в проект.",
    )
